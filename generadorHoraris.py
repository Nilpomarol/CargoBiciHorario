import streamlit as st
import pandas as pd
import openpyxl as opxl
from openpyxl.styles import NamedStyle
from datetime import datetime
import json
import os
import re


def intToHora(minutes):
    """Convert minutes into a 'hh:mm' formatted string."""
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours:02}:{mins:02}"


def horaToInt(time_str):
    """Convert a 'hh:mm' formatted string into total minutes."""
    try:
        hours, minutes = map(int, time_str.split(':'))
        return hours * 60 + minutes
    except ValueError:
        raise ValueError("Input must be in 'hh:mm' format and contain valid integers.")
    except IndexError:
        raise ValueError("Input must be in 'hh:mm' format.")


def process_routes(routes_table, time_for_delivery, pes_trike):
    """
    Processes route data, calculates arrival times, and creates a list of processed routes.

    Args:
        routes_table (str): String containing the routes data.
        time_for_delivery (int): Multiplier for delivery time.

    Returns:
        list: List of processed routes, or None if there's an error.
    """

    # Initialize an empty list to store processed route data
    processed_routes = []

    # Process each line in the input data
    for line in routes_table.splitlines():
        if not line.strip():
            continue  # Skip empty lines

        route_elements = re.split(r"\t", line)

        # Ensure the line has the expected number of elements
        parche = 0
        if len(route_elements) == 14:
            parche = -1
        elif len(route_elements) != 15:
            print(f"Warning: Unexpected line format: {line}")
            continue

        try:
            # Convert times and distances
            departure_time = horaToInt(route_elements[3 + parche])
            delivery_time = int(route_elements[6 + parche])
            multiplier = int(route_elements[12 + parche])
            arrival_time = departure_time + delivery_time + (multiplier * time_for_delivery)

            # Determine priority and bike type
            is_priority = ' w ' in route_elements[0].lower()
            bike_type = "TRIKE" if int(route_elements[9 +  parche]) <= pes_trike else "4W"
            # Create a processed route entry
            processed_route = [
                route_elements[0],  # Route identifier
                is_priority,  # Priority flag
                bike_type,# Bike type
                int(route_elements[9 + parche]),  # Weight
                route_elements[2 + parche],  # Route detail
                route_elements[10 + parche],  # Additional detail
                route_elements[3 + parche],  # Departure time
                '',  # Placeholder
                intToHora(arrival_time),  # Arrival time
                '',  # Placeholder
                delivery_time,  # Delivery time
                '',  # Placeholder
                multiplier,  # Multiplier
                '',  # Placeholder
                "",  # Placeholder
                departure_time,  # Departure time in minutes
                0  # Placeholder
            ]

            processed_routes.append(processed_route)

        except ValueError as e:
            print(f"Error processing line: {line} - {e}")
            continue

    return processed_routes


def process_workers(workers_table, week_day):
    """
    Processes worker data to extract and format relevant information based on the day of the week.

    Args:
        workers_table (str): String containing the worker data.
        week_day (int): Day of the week (0-6, where 0 is Monday).

    Returns:
        list: List of processed worker data.
    """

    # Initialize an empty list to store processed worker data
    processed_workers = []

    # Calculate the column index for the day's data
    index = week_day * 3 + 1

    # Process each line in the workers table
    for line in workers_table.splitlines():
        if not line.strip():
            continue  # Skip empty lines

        worker = re.split(r"\t", line)

        if not re.match(r'^\d+:\d+$',worker[index]):
            continue

        try:
            # Convert start time to minutes
            start_time_in_minutes = horaToInt(worker[index])

            # Convert hours worked from comma to period for float conversion
            hours_worked = float(worker[index + 2].replace(",", "."))

            # Append the processed worker data to the list
            processed_workers.append([
                worker[0],  # Worker ID or Name
                worker[index],  # Start time
                worker[index + 1],  # End time
                hours_worked,  # Hours worked
                start_time_in_minutes  # Start time in minutes
            ])
        except ValueError as e:
            print(f"Error processing line: {line} - {e}")
            continue
    processed_workers = pd.DataFrame(processed_workers, columns=["Treballador", "Entrada", "Sortida", "Hores", "Aux"])
    processed_workers = processed_workers.sort_values(by="Aux")
    processed_workers = processed_workers.drop("Aux", axis=1)
    return processed_workers


def format_stop(stop):
    """Format the stop data into a readable string."""
    return (f"id: {stop[0][8:]} de {stop[1]} a {stop[2]} "
            f"Temps d'espera (min): {stop[3]}")


def printTimeline(timeline):
    """Display the timeline for each worker in a Streamlit app, divided into three columns."""
    cols = st.columns(3)  # Create three columns
    col_index = 0

    for worker, stops in timeline.items():
        # Choose the column to write to
        with cols[col_index]:
            st.write(f"{worker}:")
            for stop in stops:
                st.write(format_stop(stop))
            st.write("")  # Add a blank line for separation
        
        # Move to the next column
        col_index = (col_index + 1) % 3


def load_data(file_path, default_data=None):
    """Load data from a JSON file."""
    if default_data is None:
        default_data = {
            "Temps Inici Torn": 12,
            "Temps Fi Torn": 5,
            "Temps Per paquet": 7,
            "Temps entre rutes": 10,
            "Marge abans - W": 10,
            "Marge despres - W": 5,
            "Marge abans - No W": 25,
            "Marge despres - No W": 15,
            "Temps maxim espera": 20,
            "Marge primera ruta torn": 10,
            "Maxim Hores Global": 9,
            "Flexibilitat +6": 0.10,
            "Flexibilitat +4": 0.20,
            "Flexibilitat -4": 0.20,
            "Pes Trike": 120
        }

    if not os.path.exists(file_path):
        print(f"File not found: {file_path}. Using default data.")
        return default_data

    try:
        with open(file_path, 'r') as file:
            return json.load(file)
    except (json.JSONDecodeError, IOError) as e:
        print(f"Error reading file {file_path}: {e}. Using default data.")
        return default_data


def save_data(data, file_path):
    """Save data to a JSON file."""
    try:
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)
        print(f"Data successfully saved to {file_path}.")
        return True
    except (IOError, TypeError) as e:
        print(f"Error saving data to {file_path}: {e}.")
        return False

def process_user_inputs():
    """Process user inputs."""
    file_path = 'variables.json'
    hipotesi = load_data(file_path)

    # List of keys from hipotesi
    keys = list(hipotesi.keys())
    
    # Create four columns in the sidebar
    col1, col2, col3, col4 = st.columns(4)
    
    # Iterate over each key and create a text input in the appropriate column
    for i, key in enumerate(keys):
        if i % 4 == 0:
            with col1:
                if key in ["Flexibilitat +6", "Flexibilitat +4", "Flexibilitat -4"]:
                    hipotesi[key] = int(st.text_input(key, int(hipotesi[key] * 100))) / 100
                else:
                    hipotesi[key] = int(st.text_input(key, value=hipotesi[key]))
        elif i % 4 == 1:
            with col2:
                if key in ["Flexibilitat +6", "Flexibilitat +4", "Flexibilitat -4"]:
                    hipotesi[key] = int(st.text_input(key, int(hipotesi[key] * 100))) / 100
                else:
                    hipotesi[key] = int(st.text_input(key, value=hipotesi[key]))
        elif i % 4 == 2:
            with col3:
                if key in ["Flexibilitat +6", "Flexibilitat +4", "Flexibilitat -4"]:
                    hipotesi[key] = int(st.text_input(key, int(hipotesi[key] * 100))) / 100
                else:
                    hipotesi[key] = int(st.text_input(key, value=hipotesi[key]))
        else:
            with col4:
                if key in ["Flexibilitat +6", "Flexibilitat +4", "Flexibilitat -4"]:
                    hipotesi[key] = int(st.text_input(key, int(hipotesi[key] * 100))) / 100
                else:
                    hipotesi[key] = int(st.text_input(key, value=hipotesi[key]))

    # Save data after collecting inputs
    save_data(hipotesi, file_path)
    return hipotesi, col4

def calculate_worker_availability(dfj, workers, database_workers, timeForDelivery, timeBetweenRoutes, globalMaxhours, timeToStartShift, timeToEndShift, earlyDepartureTimeMarginPriority, delayedDepartureTimeMarginPriority, earlyDepartureTimeMarginNoPriority, delayedDepartureTimeMarginNoPriority, maxWaitTimeBetweenRoutes, firstRouteMaxEarlyDepartureTime):
    """Calculate worker availability and route assignments."""
    pre_dft = []
    totalworkers = 0
    timeline = {}
    dfj_general = []

    
    trikesInHub = {}
    fourWheelsInHub = {}


    dfj_hub = dfj.groupby("Hub")
    for hub, dfj in dfj_hub:
        #sort by delivery exit time
        dfj = dfj.sort_values(by="order")
        trikeDisponibility = {}
        fourWheelsDisponibility = {}
        numberTrikes = 0
        number4Wheels = 0
        #sorts entries by Hub and Start Time

        for index, row in dfj.iterrows():

            #worker asigned to the route
            asignedTo = -1
            bikeType = row["Tipus Bici"]
            #expected initial time of the route
            expectedInitialTime = horaToInt(row["Hora Inici Ruta Plnif"])
            
            if row["Prioritari"]:
                maxDelayedInitialTime = expectedInitialTime + delayedDepartureTimeMarginPriority #maximum late time to start the route
                maxEarlyInitialTime = expectedInitialTime - earlyDepartureTimeMarginPriority #maximum early time to start the route
            else:
                maxDelayedInitialTime = expectedInitialTime + delayedDepartureTimeMarginNoPriority  #maximum late time to start the route with added non-priority margin
                maxEarlyInitialTime = expectedInitialTime - earlyDepartureTimeMarginNoPriority #maximum early time to start the route with added non-priority margin
                
            #time to complete the route
            timeToCompleteRoute = row["Temps Recorregut Ruta"] + row["Num Entregues"] * timeForDelivery
            dfj.at[index, "Temps Total Ruta"] = timeToCompleteRoute
            #end time of the route
            endTime = horaToInt(row["Hora Fi Ruta"]) + timeBetweenRoutes

            for t, value in workers.items(): #check if any worker is available   

                available = value[0] < maxDelayedInitialTime #check if worker has ended route before max delayed time
                
                if t in database_workers:
                    hoursLeftShift = (pre_dft[value[1]][4] + timeToCompleteRoute/60) <= min(database_workers[t][1], globalMaxhours) #Hours left to the shift
                else: 
                    hoursLeftShift = (pre_dft[value[1]][4] + timeToCompleteRoute/60) <= globalMaxhours

                exceededMaxWaitTime = (maxEarlyInitialTime - value[0]) < maxWaitTimeBetweenRoutes #worker waiting without a route

                correctHub = workers[t][2] == row['Hub']

                if available and hoursLeftShift and exceededMaxWaitTime and correctHub:

                    #assign the job to worker t, and update the corresponding data structures
                    asignedTo = t

                    routeStartTime = max(value[0], maxEarlyInitialTime) #Start time of the route

                    dfj.at[index, "Hora Inici Ruta Real"] = intToHora(routeStartTime) #update the table with the actual start time of the route
                    
                    endTime = routeStartTime + timeToCompleteRoute + timeBetweenRoutes #time when the worker can start the next route

                    dfj.at[index, "Hora Fi Ruta"] = intToHora(routeStartTime + timeToCompleteRoute) #update the table with the actual end time of the route
                    dfj.at[index, "Plnif vs Real Min"] = -(horaToInt(row["Hora Inici Ruta Plnif"]) - routeStartTime) #difference of starting time between plan and actual
                    dfj.at[index, "Inici Seguent Ruta"] = intToHora(endTime) #time at which the worker that did this route is available for the next one

                    workers[t] = (endTime, value[1], value[2])
                    
                    pre_dft[value[1]][3] = intToHora(routeStartTime + timeToCompleteRoute + timeToEndShift) #update the provisional end of the shift
                    pre_dft[value[1]][4] = round((horaToInt(pre_dft[value[1]][3]) - horaToInt(pre_dft[value[1]][2]))/60,1) #update the total hours worked
                    
                    waitingTime = routeStartTime-(horaToInt(timeline[asignedTo][-1][2])+10)
                    timeline[asignedTo][-1] = (timeline[asignedTo][-1][0], timeline[asignedTo][-1][1], intToHora(horaToInt(timeline[asignedTo][-1][2])+10), timeline[asignedTo][-1][3])
                    timeline[asignedTo].append((row["Id"].split()[0],intToHora(routeStartTime), intToHora(routeStartTime + timeToCompleteRoute), waitingTime))
                    
                    number4Wheels, numberTrikes, fourWheelsDisponibility, trikeDisponibility = updateBikeAssignment(timeBetweenRoutes, numberTrikes, number4Wheels, trikeDisponibility, fourWheelsDisponibility, row, bikeType, endTime, routeStartTime)

                    break
                    
                    

            if asignedTo == -1: #No worker is available, then, add another worker
                
                maxEarlyInitialTime = expectedInitialTime - firstRouteMaxEarlyDepartureTime #Start time of the route

                dfj.at[index, "Hora Inici Ruta Real"] = intToHora(maxEarlyInitialTime) #update the table with the actual start time of the route

                endTime = maxEarlyInitialTime + timeToCompleteRoute + timeBetweenRoutes #time when the worker can start the next route
                
                dfj.at[index, "Hora Fi Ruta"] = intToHora(maxEarlyInitialTime + timeToCompleteRoute)
                dfj.at[index, "Plnif vs Real Min"] = -(horaToInt(row["Hora Inici Ruta Plnif"]) - maxEarlyInitialTime) #difference of starting time between plan and actual
                dfj.at[index, "Inici Seguent Ruta"] = intToHora(endTime) #time at which the worker that did this route is available for the next one

                id = totalworkers #New worker assigned to this route
                workers[id] = (endTime, len(pre_dft), row['Hub']) #add it to the dict with the active workers and their last route end time

                startShift= maxEarlyInitialTime - timeToStartShift
                #time it would end the shift if no more routes would be done
                provisionalEndShift = maxEarlyInitialTime + timeToCompleteRoute + timeToEndShift
                provisionalShiftHours = round((provisionalEndShift - startShift)/60,1)

                newWorker = [row["Hub"], id, intToHora(startShift), intToHora(provisionalEndShift), provisionalShiftHours]
                
                pre_dft.append(newWorker) #add worker to the database

                timeline[id] = [(row["Id"].split()[0],intToHora(maxEarlyInitialTime), intToHora(maxEarlyInitialTime + timeToCompleteRoute), "")]

                asignedTo = id
                totalworkers += 1

                number4Wheels, numberTrikes, fourWheelsDisponibility, trikeDisponibility = updateBikeAssignment(timeBetweenRoutes, numberTrikes, number4Wheels, trikeDisponibility, fourWheelsDisponibility, row, bikeType, endTime, maxEarlyInitialTime)

            dfj.at[index, "Assignacio Prov"] = asignedTo
        #add the modified dataframe with the assignments to the list of dataframes
        dfj_general.append(dfj)
        trikesInHub[hub] = numberTrikes
        fourWheelsInHub[hub] = number4Wheels

    dft = pd.DataFrame(pre_dft, columns=['Hub', 'worker', 'Hora Inici Torn', 'Hora Final Torn', 'Hores Totals'])

    dft = dft.sort_values(by='Hores Totals', ascending=False)
    
    idToWorker = {}
    i = 0


    #Assign workers to the shift the best fits their hours

    dft.insert(1, "Treballador", '')
    extraWorkers = 65

    for index, row in dft.iterrows():
        if i in database_workers:
            idToWorker[row["worker"]] = database_workers[i][0]
        else:
            idToWorker[row["worker"]] = chr(extraWorkers)
            extraWorkers += 1
        dft.at[index, "Treballador"] = idToWorker[row["worker"]]

        i += 1
    
    dft = dft.drop("worker", axis=1)

    #Add the names of the workers to the assignments
    dfj = pd.concat(dfj_general)
    dfj = dfj.drop("order", axis=1)
    i = 0

    for index, row in dfj.iterrows():
        dfj.at[index, "Assignació"] = idToWorker[row["Assignacio Prov"]]
        if row["Assignacio Prov"] in timeline:
            timeline[idToWorker[row["Assignacio Prov"]]] = timeline.pop(row["Assignacio Prov"])

    dfj = dfj.drop("Assignacio Prov", axis=1)
    
    return dfj, dft, timeline, trikesInHub, fourWheelsInHub

def updateBikeAssignment(timeBetweenRoutes, numberTrikes, number4Wheels, trikeDisponibility, fourWheelsDisponibility, row, bikeType, endTime, routeStartTime):
    asignedbike = False
    
    if bikeType == "TRIKE":
        for index, trike in trikeDisponibility.items():
            if trike[0] <= routeStartTime:
                trike[0] = endTime - timeBetweenRoutes
                asignedbike = True
                break
        if not asignedbike:
            trikeDisponibility[numberTrikes] = [endTime - timeBetweenRoutes, row['Hub']]
            numberTrikes += 1
    else:
        for index, fourWheels in fourWheelsDisponibility.items():
            if fourWheels[0] <= routeStartTime:
                fourWheels[0] = endTime - timeBetweenRoutes
                asignedbike = True
                break
        if not asignedbike:
            fourWheelsDisponibility[number4Wheels] = [endTime - timeBetweenRoutes, row['Hub']]
            number4Wheels += 1
    return number4Wheels, numberTrikes, fourWheelsDisponibility, trikeDisponibility

def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        
        # Find the maximum length of the content in each cell
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
        # Set the column width
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


def generate_excel_file(dfj, dft, workers_sants, workers_napols, additionalInfoList, hipotesi, timeline, workerList, numberTrikes, number4Wheels):
    """Generate and format the Excel file."""
    wb = opxl.Workbook()
    wb.save('output.xlsx')
    dfj = dfj.drop("Prioritari", axis=1)
    with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
        dfj.to_excel(writer, sheet_name='Taula General', startcol=1, startrow=1)

        workerList = {}

        dfj_group = dfj.groupby('Hub')
        dft_group = dft.groupby('Hub')

        for hub, dft_hub in dft_group:
            dft_hub = dft_hub.sort_values(by='Hora Inici Torn')
            dft_hub.index = list(range(1, len(dft_hub) + 1))
            dft_hub.to_excel(writer, sheet_name=hub, startcol=1, startrow=1)

            if hub == "Sants" and workers_sants is not None:
                workers_sants.to_excel(writer, sheet_name=hub, startcol=8, startrow=1, index=False)
            elif workers_napols is not None:
                workers_napols.to_excel(writer, sheet_name=hub, startcol=8, startrow=1, index=False)

            workerListAux = []
            rowToWrite = len(dft_hub) + 9 + 7
            dfj_hub = dfj_group.get_group(hub)
            dfj_hub.index = list(range(1, len(dfj_hub) + 1))
            dfj_hub.to_excel(writer, sheet_name=hub, startcol=1, startrow=rowToWrite)
            rowToWrite += len(dfj_hub) + 5
            colToWrite = 1

            for index, row in dft_hub.iterrows():
                workerTimeline = timeline[row["Treballador"]]
                df_workerTimeline = pd.DataFrame(workerTimeline, columns=['ID', 'Inici Ruta', 'Fi Ruta', 'Temps Espera Min'])
                df_workerTimeline.to_excel(writer, sheet_name=hub, startcol=colToWrite, startrow=rowToWrite)
                workerListAux.append((row["Treballador"], rowToWrite))
                colToWrite += 7

                if colToWrite == 22:
                    colToWrite = 1
                    rowToWrite += len(df_workerTimeline) + 8

            workerList[hub] = workerListAux

    try:
        wb = opxl.load_workbook('output.xlsx')
        # Create a NamedStyle for the percentage format
        percentage_style = NamedStyle(name="percentage_style", number_format='0.00%')

    except Exception as e:
        print(f"Error loading the Excel file: {e}")
        wb = opxl.Workbook()
        wb.save('output_temp.xlsx')
        wb = opxl.load_workbook('output_temp.xlsx')

    for data in additionalInfoList:
        hub = data[0]
        if hub in wb.sheetnames:
            sheet = wb[hub]
        else:
            print(f"Skipping {hub}: Worksheet does not exist.")
            continue

        row_number = data[3] + 2
        column_letter = 'G'
        formula = f"=SUM({column_letter}3:{column_letter}{row_number})"
        row_number += 3
        sheet.cell(row=row_number, column=5).value = "Total Hores"
        sheet.cell(row=row_number, column=6).value = formula
        sheet.cell(row=row_number, column=11).value = "Total Hores"
        sheet.cell(row=row_number, column=12).value = f"=SUM(L3:L{row_number-1})"
        row_number += 1
        sheet.cell(row=row_number, column=5).value = "Num treballadors"
        sheet.cell(row=row_number, column=6).value = data[3]
        row_number += 1
        sheet.cell(row=row_number, column=5).value = "Num Rutes"
        sheet.cell(row=row_number, column=6).value = data[1]
        row_number += 1
        sheet.cell(row=row_number, column=5).value = "Total Paquets"
        sheet.cell(row=row_number, column=6).value = data[4]
        row_number += -1
        sheet.cell(row=row_number, column=9).value = "TRIKES"
        sheet.cell(row=row_number, column=10).value = data[5]
        sheet.cell(row=row_number, column=11).value = f"=J{row_number}/F{row_number}"
        sheet.cell(row=row_number, column=11).style = percentage_style   
        row_number += 1
        sheet.cell(row=row_number, column=9).value = "4W"
        sheet.cell(row=row_number, column=10).value = data[1] - data[5]
        sheet.cell(row=row_number, column=11).value = f"=J{row_number}/F{row_number-1}"
        sheet.cell(row=row_number, column=11).style = percentage_style
        row_number += 1
        sheet.cell(row=row_number, column=9).value = "TRIKES Min"
        sheet.cell(row=row_number, column=10).value = numberTrikes[hub]
        row_number += 1
        sheet.cell(row=row_number, column=9).value = "4w  Min"
        sheet.cell(row=row_number, column=10).value = number4Wheels[hub]
        row_number += 1

        cellRangeString = 'O' + str(row_number + 5) + ':O' + str(row_number + 5 + len(dfj_group.get_group(data[0])))
        cellRange = sheet[cellRangeString]

        for row in cellRange:
            for cell in row:
                if cell.value is not None:
                    try:
                        numeric_value = int(cell.value)
                        if numeric_value > 0:
                            cell.font = opxl.styles.Font(color='FF0000')
                    except ValueError:
                        pass

        colToWrite = 1
        for worker_tuple in workerList.get(data[0], []):
            row_number = worker_tuple[1]
            sheet.cell(row=row_number + 1, column=colToWrite + 1).value = worker_tuple[0]
            colToWrite += 7
            if colToWrite == 22:
                colToWrite = 1

    sheet = wb['Taula General']
    cellRangeString = 'O3:O' + str(len(dfj) + 3)
    cellRange = sheet[cellRangeString]

    for row in cellRange:
        for cell in row:
            if cell.value is not None:
                try:
                    numeric_value = int(cell.value)
                    if numeric_value > 0:
                        cell.font = opxl.styles.Font(color='FF0000')
                except ValueError:
                    pass

    sheet = wb.create_sheet("Hipotesi")
    row_number = 2
    for key in hipotesi:
        sheet.cell(row=row_number, column=2).value = key
        sheet.cell(row=row_number, column=3).value = hipotesi[key]
        row_number += 1
        
    # Loop through each sheet in the workbook and adjust sizes
    for sheet in wb.worksheets:
        adjust_column_widths(sheet)


    wb.save('output.xlsx')



def display_ui(dfj, dft, timeline, numberTrikes, number4Wheels):
    """Display the Streamlit UI components."""
    col3, cols = st.columns(2)
    dft_grouped = dft.groupby("Hub")
    dfj = dfj.drop("Prioritari", axis=1)
    additional_info_list = []
    for hub, dft_hub in dft_grouped:

        # Ordenem el DataFrame per la columna 'Hora Inici Torn'
        # dft_hub['Hora Inici Torn'] = pd.to_datetime(dft_hub['Hora Inici Torn'], format='%H:%M').sort_values().dt.strftime('%H:%M')
        dft_hub = dft_hub.sort_values(by="Hora Inici Torn")
        # Mostrem el DataFrame ordenat
        dft_hub.index = range(1, len(dft_hub) + 1)
        total_hours = dft_hub["Hores Totals"].sum()
        workers_in_hub = len(dft_hub)
        number_deliveries = len(dfj[dfj["Hub"] == hub])
        number_packages = dfj[dfj["Hub"] == hub]["Num Entregues"].sum()
        trike_bikes = len(dfj[(dfj["Hub"] == hub) & (dfj["Tipus Bici"] == "TRIKE")])

        additional_info_list.append((hub, number_deliveries, total_hours, workers_in_hub, number_packages, trike_bikes))
        
        # Display information in columns
        with (col3 if hub == "Sants" else cols):
            st.write(f"Informació per al Hub: {hub}")
            st.write(f"TRIKES: {numberTrikes[hub]}")
            st.write(f"4W: {number4Wheels[hub]}")
            col5, col6, col7, col8 = st.columns(4)
            with col5:
                st.write(f"Hores Totals: {total_hours:.1f}")
            with col6:
                st.write(f"Numero de treballadors: {workers_in_hub}")
            with col7:
                st.write(f"Numero de rutes: {number_deliveries}")
            with col8:
                st.write(f"Paquets Totals: {number_packages}")
             
            st.write(f"Taula amb les hores del empleats per al Hub: {hub}")
            st.write(dft_hub)

    st.write("Taula amb la assignació de treballs")
    st.write(dfj)

    # Print timeline
    printTimeline(timeline)

    # Create and display the download button
    
    return additional_info_list

def executarGenerarHoraris():
    
    hipotesi, col4 = process_user_inputs()
    routes_table = st.text_area("taula amb les rutes")
    colsants, colnapols = st.columns(2)
    with colsants:
        workers_sants_table = st.text_area("HORARIS SANTS")
    with colnapols:
        workers_napols_table = st.text_area("HORARIS NAPOLS")

    if routes_table:
        dfj_hub = pd.DataFrame(process_routes(routes_table, hipotesi["Temps Per paquet"], hipotesi["Pes Trike"]), columns=["Id", "Prioritari", "Tipus Bici", "Pes", "Data", "Hub", "Hora Inici Ruta Plnif",
                                "Hora Inici Ruta Real", "Hora Fi Ruta", "Inici Seguent Ruta",
                                "Temps Recorregut Ruta", "Temps Total Ruta", "Num Entregues", "Assignació", "Assignacio Prov", "order", "Plnif vs Real Min"])
        
        weekday = datetime.strptime(dfj_hub["Data"].iloc[0], "%d/%m/%Y").weekday()

        workers_sants = process_workers(workers_sants_table, weekday) if workers_sants_table else None
        workers_napols = process_workers(workers_napols_table, weekday) if workers_napols_table else None

        dfj, dft, timeline,numberTrikes, number4Wheels = calculate_worker_availability(
            dfj_hub,
            {},  # Replace with actual worker data
            {},  # Replace with actual database worker data
            hipotesi["Temps Per paquet"],
            hipotesi["Temps entre rutes"],
            hipotesi["Maxim Hores Global"],
            hipotesi["Temps Inici Torn"],
            hipotesi["Temps Fi Torn"],
            hipotesi["Marge abans - W"],
            hipotesi["Marge despres - W"],
            hipotesi["Marge abans - No W"],
            hipotesi["Marge despres - No W"],
            hipotesi["Temps maxim espera"],
            hipotesi["Marge primera ruta torn"]
        )

        
        additional_info_list = display_ui(dfj, dft, timeline, numberTrikes, number4Wheels)
        generate_excel_file(dfj, dft, workers_sants, workers_napols, additional_info_list, hipotesi, timeline, {}, numberTrikes, number4Wheels)

        with col4:
            st.write("")

            with open("output.xlsx", "rb") as file:
                st.download_button(
                    label='Descarregar Fitxer Excel',
                    data=file.read(),
                    file_name='output.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
        


if __name__ == "__main__":
    executarGenerarHoraris()
